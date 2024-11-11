import uuid, os, webcolors, traceback, re, json, calendar, cv2, numpy as np, pandas as pd, io, xlsxwriter, pytz
from datetime import timedelta, datetime
from dateutil.relativedelta import relativedelta
from io import BytesIO
from sklearn.cluster import KMeans
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model, update_session_auth_hash, authenticate, login
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.mail import send_mail
from django.conf import settings
from django.forms import modelformset_factory
from django.db.models import Sum, F, Count
from django.db.models.functions import TruncHour, TruncDate
from django.utils import timezone
from django.utils.timezone import make_aware
from mainapp.decorators import user_type_required
from mainapp.utils import hash_url
from mainapp.models import Order, OrderItem
from adminapp.models import Brand, Category, WatchType, Collection, Material, Feature, BaseWatch, WatchDetails, WatchMaterials, WatchImage, BrandApproval
from .forms import (
    VendorAddressForm, VendorProfileEditForm, VendorProfileForm, 
    VendorRegistrationFormStep1, VendorRegistrationFormStep2, 
    BaseWatchForm, BrandSelectionForm, BrandApprovalRequestForm, 
    WatchDetailsForm, WatchMaterialsForm, WatchImageFormSet, 
    VendorOnboardingForm, BaseWatchUpdateForm, WatchDetailsUpdateForm, 
    WatchMaterialsUpdateForm, WatchImageUpdateFormSet
)
from .models import VendorAddress, VendorProfile

User = get_user_model()
@never_cache
def home(request):
    return render(request,'vendorapp/home.html')

@never_cache
def vendor_register_step1(request):
    if request.user.is_authenticated and request.user.role == 'vendor':
        return redirect('vendorapp:index')
    
    if request.method == 'POST':
        user_form = VendorRegistrationFormStep1(request.POST)
        if user_form.is_valid():
            # Store form data in session
            request.session['vendor_user_data'] = user_form.cleaned_data
            return redirect('vendorapp:vendor_register_step2')
    else:
        user_form = VendorRegistrationFormStep1()
    
    # Always pass the form to the template, whether it's a new form or one with errors
    return render(request, 'vendorapp/register_step1.html', {'user_form': user_form})

@never_cache
def vendor_register_step2(request):
    if 'vendor_user_data' not in request.session:
        return redirect('vendorapp:vendor_register_step1')

    if request.method == 'POST':
        profile_form = VendorProfileForm(request.POST)
        password_form = VendorRegistrationFormStep2(request.POST)
        if profile_form.is_valid() and password_form.is_valid():
            user_data = request.session['vendor_user_data']
            user = User.objects.create_user(
                fullname=user_data['fullname'],
                email=user_data['email'],
                username=user_data['email'],
                role='vendor',
                password=password_form.cleaned_data['password1']
            )
            user.set_password(password_form.cleaned_data['password1'])
            user.save()

            vendor_profile = profile_form.save(commit=False)
            vendor_profile.user = user
            vendor_profile.contact_email = user_data['contact_email']
            vendor_profile.gst_number = user_data['gst_number']
            vendor_profile.contact_phone = user_data['contact_phone']
            vendor_profile.save()

            # Clear session data
            del request.session['vendor_user_data']

            # Authenticate and login the user
            user = authenticate(request, username=user_data['email'], email=user_data['email'], password=password_form.cleaned_data['password1'])
            login(request, user)

            messages.success(request, "You have successfully registered as a vendor.")
            return redirect('vendorapp:vendor_onboarding')
        else:
            print(profile_form.errors)  # Debug: Print form errors if the form is not valid
            print(password_form.errors)  # Debug: Print form errors if the form is not valid
    else:
        profile_form = VendorProfileForm()
        password_form = VendorRegistrationFormStep2()
    return render(request, 'vendorapp/register_step2.html', {'profile_form': profile_form, 'password_form': password_form})

@never_cache
@login_required
@user_type_required('vendor')
def vendor_onboarding(request):
    vendor_profile = request.user.vendorprofile
    form = VendorOnboardingForm(instance=vendor_profile)
    
    if vendor_profile.user.is_verified:
        if request.method == 'POST':
            form = VendorOnboardingForm(request.POST, instance=vendor_profile)
            if form.is_valid():
                form.save()
                vendor_profile.is_onboarding_completed = True
                vendor_profile.save()
                messages.success(request, "Your profile has been updated successfully.")
                return redirect('vendorapp:index')
        else:
            form = VendorOnboardingForm(instance=vendor_profile)
    else:
        messages.error(request, "Please verify your email address before proceeding.")
    
    context = {
        'form': form,
        'vendor_profile': vendor_profile,
    }
    return render(request, 'vendorapp/vendor_onboarding.html', context)

@never_cache
@login_required
@user_type_required('vendor')
def vendor_send_verification_email(request):
    user = request.user
    if user.is_verified:
        return JsonResponse({"success": False, "message": "Your email is already verified."})

    # Generate a new verification token
    user.email_verification_token = uuid.uuid4()
    user.save()

    verification_url = request.build_absolute_uri(
        reverse('vendorapp:vendor_verify_email', kwargs={'token': user.email_verification_token})
    )

    subject = 'Verify your email address'
    message = f'Please click on the link below to verify your email address:\n\n{verification_url}'
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = [user.email]

    try:
        send_mail(subject, message, from_email, recipient_list)
        return JsonResponse({"success": True, "message": "Verification email sent. Please check your inbox."})
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Failed to send email: {str(e)}"})
    
@never_cache
def vendor_verify_email(request, token):
    try:
        user = User.objects.get(email_verification_token=token, role='vendor')
        user.is_verified = True
        user.email_verification_token = None
        user.save()
        messages.success(request, "Your email has been successfully verified.")
    except User.DoesNotExist:
        messages.error(request, "Invalid verification link.")

    return redirect('vendorapp:vendor_login')

@never_cache
@login_required
@user_type_required('vendor')
def vendor_check_email_status(request):
    user = request.user
    return JsonResponse({"is_verified": user.is_verified})

def validate_company_name(request):
    company_name = request.GET.get('company_name', None)
    if company_name:
        company_name = company_name.lower()
        data = {
            'is_taken': VendorProfile.objects.filter(company_name__iexact=company_name.lower()).exists()
        }
    else:
        data = {'is_taken': False}
    return JsonResponse(data)

@never_cache
def vendor_login(request):
    if request.user.is_authenticated and request.user.role == 'vendor':
        return redirect('vendorapp:index')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        print(email, password)
        user = authenticate(request, username=email, email=email, password=password)
        
        if user is not None and user.role == 'vendor':
            login(request, user)
            messages.success(request, 'You have successfully logged in.')
            return redirect('vendorapp:index')
        else:
            messages.error(request, 'Invalid email or password.')
    
    return render(request, 'vendorapp/vendor_login.html')
@never_cache
@login_required
@user_type_required('vendor')
def index(request):
    vendor_profile = VendorProfile.objects.get(user=request.user)
    if vendor_profile.is_onboarding_completed:
        if vendor_profile.approval_status == 'Approved':
            # Get total products
            total_products = BaseWatch.objects.filter(vendor=vendor_profile).count()
            
            # Get new products added in the last week
            one_week_ago = timezone.now() - timedelta(days=7)
            # new_products_this_week = BaseWatch.objects.filter(vendor=vendor_profile, created_at__gte=one_week_ago).count()
            new_products_this_week = BaseWatch.objects.filter(vendor=vendor_profile).count()
            product_growth_percentage = (new_products_this_week / total_products * 100) if total_products > 0 else 0
            
            # Get total sales for the last 30 days
            thirty_days_ago = timezone.now() - timedelta(days=30)
            total_sales = BaseWatch.objects.filter(
                vendor=vendor_profile,
                orderitem__order__created_at__gte=thirty_days_ago
            ).aggregate(
                total_sales=Sum('orderitem__quantity')
            )['total_sales'] or 0
            
            # Get sales growth percentage
            previous_thirty_days = timezone.now() - timedelta(days=60)
            previous_sales = BaseWatch.objects.filter(
                vendor=vendor_profile,
                orderitem__order__created_at__gte=previous_thirty_days,
                orderitem__order__created_at__lt=thirty_days_ago
            ).aggregate(
                total_sales=Sum('orderitem__quantity')
            )['total_sales'] or 0
            
            sales_growth_percentage = ((total_sales - previous_sales) / previous_sales * 100) if previous_sales > 0 else 0
            
            # Get total orders
            total_orders = BaseWatch.objects.filter(
                vendor=vendor_profile,
                orderitem__isnull=False
            ).aggregate(
                total_orders=Count('orderitem__order', distinct=True)
            )['total_orders'] or 0
            
            # Get order growth percentage
            new_orders_this_week = BaseWatch.objects.filter(
                vendor=vendor_profile,
                orderitem__order__created_at__gte=one_week_ago
            ).aggregate(
                new_orders=Count('orderitem__order', distinct=True)
            )['new_orders'] or 0
            
            order_growth_percentage = (new_orders_this_week / total_orders * 100) if total_orders > 0 else 0
            
            # Get 4 most recent orders
            recent_orders = Order.objects.filter(
                items__watch__vendor=vendor_profile
            ).distinct().order_by('-created_at')[:4]

            context = {
                'total_sales': total_sales,
                'sales_growth_percentage': round(sales_growth_percentage, 1),
                'total_products': total_products,
                'product_growth_percentage': round(product_growth_percentage, 1),
                'total_orders': total_orders,
                'order_growth_percentage': round(order_growth_percentage, 1),
                'recent_orders': recent_orders,
            }
            
            return render(request, 'vendorapp/index.html', context)
        else:
            return render(request, 'vendorapp/vendor_application_pending.html')
    else:
        return redirect('vendorapp:vendor_onboarding')
    
@never_cache
@login_required
@user_type_required('vendor')
def vendor_profile(request):
    vendor = request.user.vendorprofile
    vendor_address = VendorAddress.objects.filter(vendor=vendor).first()
    
    if request.method == 'POST':
        form = VendorProfileEditForm(request.POST, instance=vendor)
        address_form = VendorAddressForm(request.POST, instance=vendor_address)
        if form.is_valid() and address_form.is_valid():
            form.save()
            address_form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('vendorapp:vendor_profile')
    else:
        form = VendorProfileEditForm(instance=vendor)
        address_form = VendorAddressForm(instance=vendor_address)

    context = {
        'vendor': vendor,
        'vendor_address': vendor_address,
        'form': form,
        'address_form': address_form,
    }
    return render(request, 'vendorapp/vendor_profile.html', context)

@never_cache
@login_required
@user_type_required('vendor')
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            # Additional custom validation
            new_password = form.cleaned_data.get('new_password1')
            if len(new_password) < 8:
                form.add_error('new_password1', 'Password must be at least 8 characters long.')
            elif not re.search(r'[A-Z]', new_password):
                form.add_error('new_password1', 'Password must contain at least one uppercase letter.')
            elif not re.search(r'[a-z]', new_password):
                form.add_error('new_password1', 'Password must contain at least one lowercase letter.')
            elif not re.search(r'\d', new_password):
                form.add_error('new_password1', 'Password must contain at least one number.')
            
            if not form.errors:
                user = form.save()
                update_session_auth_hash(request, user)  # Important!
                messages.success(request, 'Your password was successfully updated!')
                return JsonResponse({'success': True, 'message': 'Your password was successfully updated!'})
        
        # If form is not valid or custom validation failed
        errors = {field: form.errors[field] for field in form.errors}
        return JsonResponse({'success': False, 'errors': errors})
    
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'vendorapp/change_password.html', {'form': form})


@never_cache
@login_required
@user_type_required('vendor')
def check_unique_model_name(request):
    model_name = request.GET.get('model_name', '')
    model_name = model_name.replace('%20', ' ')
    is_unique = not BaseWatch.objects.filter(model_name=model_name).exists()
    return JsonResponse({'is_unique': is_unique})

@never_cache
@login_required
@user_type_required('vendor')
def product_list(request):
    vendor_profile = VendorProfile.objects.get(user=request.user)
    products = BaseWatch.objects.filter(vendor=vendor_profile).select_related('brand', 'category', 'watch_type')
    form = BaseWatchForm()
    form_fields = {field: str(form[field]) for field in form.fields}
    watch_types = WatchType.objects.all()
    return render(request, 'vendorapp/product_list.html', {
        'products': products,
        'form_fields': form_fields,
        'watch_types': watch_types
    })


@require_POST
@login_required
@user_type_required('vendor')
def toggle_product_status(request):
    product_id = request.POST.get('product_id')
    status = request.POST.get('status')
    
    try:
        product = BaseWatch.objects.get(id=product_id, vendor=request.user.vendorprofile)
        
        # Convert status to boolean
        new_status = status.lower() == 'true'
        
        # Toggle the status
        product.is_active = new_status
        product.save()
        
        return JsonResponse({'success': True, 'new_status': product.is_active})
    except BaseWatch.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@never_cache
@login_required
@user_type_required('vendor')
def delete_product(request, product_id):
    product = get_object_or_404(BaseWatch, id=product_id)
    product.delete()
    messages.success(request, 'Product deleted successfully.')
    return redirect('vendorapp:product_list')

@never_cache
@login_required
@user_type_required('vendor')
def edit_product(request, product_id):
    product = get_object_or_404(BaseWatch, id=product_id, vendor=request.user.vendorprofile)
    
    if request.method == 'POST':
        base_watch_form = BaseWatchUpdateForm(request.POST, request.FILES, instance=product)
        details_form = WatchDetailsUpdateForm(request.POST, instance=product.details)
        materials_form = WatchMaterialsUpdateForm(request.POST, instance=product.materials)
        
        if all([base_watch_form.is_valid(), details_form.is_valid(), materials_form.is_valid()]):
            try:
                base_watch = base_watch_form.save()
                details_form.save()
                materials_form.save()
                
                # Handle primary image
                if 'primary_image' in request.FILES:
                    base_watch.primary_image = request.FILES['primary_image']
                    # base_watch.save()
                    # Extract dominant color
                    dominant_color = get_dominant_color(base_watch.primary_image.path)
                    base_watch.dominant_color = dominant_color
                    base_watch.save()
                
                # Handle deletion of existing images
                images_to_delete = request.POST.getlist('delete_images')
                WatchImage.objects.filter(id__in=images_to_delete).delete()
                
                # Handle new additional images
                new_images = request.FILES.getlist('additional_images')
                for image in new_images:
                    WatchImage.objects.create(base_watch=base_watch, image=image)
                
                messages.success(request, f"Product {base_watch.model_name} has been updated successfully.")
                return redirect('vendorapp:product_list')
            except Exception as e:
                messages.error(request, f"An error occurred while saving the product: {str(e)}")
        else:
            print(base_watch_form.errors)
            print(details_form.errors)
            print(materials_form.errors)
            messages.error(request, "There were errors in the form. Please check below for details.")
    else:
        base_watch_form = BaseWatchUpdateForm(instance=product)
        details_form = WatchDetailsUpdateForm(instance=product.details)
        materials_form = WatchMaterialsUpdateForm(instance=product.materials)
    
    context = {
        'base_watch_form': base_watch_form,
        'details_form': details_form,
        'materials_form': materials_form,
        'product': product,
    }
    return render(request, 'vendorapp/edit_product.html', context)

@never_cache
@login_required
@user_type_required('vendor')
def delete_product_image(request, image_id):
    image = get_object_or_404(WatchImage, id=image_id)
    if image.base_watch.vendor == request.user.vendorprofile:
        image.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=403)

@never_cache
@login_required
@user_type_required('vendor')
def manage_stock(request):
    vendor_profile = request.user.vendorprofile
    products = BaseWatch.objects.filter(vendor=vendor_profile)
    return render(request, 'vendorapp/manage_stock.html', {'products': products})

@never_cache
@login_required
@user_type_required('vendor')
def update_stock(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(BaseWatch, id=product_id, vendor=request.user.vendorprofile)
        stock_change = request.POST.get('stock_change')
        try:
            stock_change = int(stock_change)
            new_available_stock = product.available_stock + stock_change
            if new_available_stock >= 0:
                product.available_stock = new_available_stock
                if stock_change > 0:
                    # Increasing stock
                    product.total_stock += stock_change
                else:
                    # Decreasing stock
                    product.total_stock = max(product.total_stock + stock_change, product.available_stock)
                
                product.is_in_stock = product.available_stock > 0
                product.save()
                return JsonResponse({
                    'success': True, 
                    'available_stock': product.available_stock,
                    'total_stock': product.total_stock,
                    'sold_stock': product.total_stock - product.available_stock,
                    'is_in_stock': product.is_in_stock
                })
            else:
                return JsonResponse({'success': False, 'error': 'Available stock cannot be negative'})
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid stock change value'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@never_cache
@login_required
@user_type_required('vendor')
def order_list(request):
    vendor_profile = request.user.vendorprofile
    orders = Order.objects.filter(
        items__watch__vendor=vendor_profile
    ).distinct().order_by('-created_at')

    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        orders = orders.filter(created_at__gte=start_date)
    if end_date:
        orders = orders.filter(created_at__lte=end_date)

    paginator = Paginator(orders, 10)  # Show 10 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }
    return render(request, 'vendorapp/order_list.html', context)

@never_cache
@login_required
@user_type_required('vendor')
def download_orders(request):
    vendor_profile = request.user.vendorprofile
    orders = Order.objects.filter(
        items__watch__vendor=vendor_profile
    ).distinct().order_by('-created_at')

    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        orders = orders.filter(created_at__gte=start_date)
    if end_date:
        orders = orders.filter(created_at__lte=end_date)

    # Prepare data for Excel
    data = []
    for order in orders:
        for item in order.items.filter(watch__vendor=vendor_profile):
            data.append({
                'Order ID': order.order_id,
                'Model Name': item.watch.model_name,
                'Customer': order.user.email,
                'Total Amount': item.price * item.quantity,
                'Status': order.get_status_display(),
                'Date': order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Quantity': item.quantity,
            })

    # Create DataFrame and Excel file
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Orders')

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
    return response


@never_cache
@login_required
@user_type_required('vendor')
def order_detail(request, order_id):
    # Get order with related data in a single query
    order = get_object_or_404(
        Order.objects.select_related(
            'user',
            'user__profile',
            'address'
        ).prefetch_related(
            'items',
            'items__watch'
        ),
        order_id=order_id
    )
    
    context = {
        'order': order,
        'title': f'Order #{order.order_id}',
    }
    
    return render(request, 'vendorapp/order_detail.html', context)
# test

@login_required
@user_type_required('vendor')
def add_product_step1(request):
    vendor = request.user.vendorprofile
    brands = Brand.objects.all()
    brand_approvals = BrandApproval.objects.filter(vendor=vendor)

    if request.method == 'POST':
        selected_brand_id = request.POST.get('brand')
        listing_type = request.POST.get('listing_type')
        if selected_brand_id and listing_type:
            brand = get_object_or_404(Brand, id=selected_brand_id)
            approval = brand_approvals.filter(brand=brand).first()
            
            # Store brand_id and listing_type in session
            request.session['selected_brand_id'] = selected_brand_id
            request.session['listing_type'] = listing_type
            
            if approval and approval.is_approved:
                if listing_type == 'single':
                    return redirect('vendorapp:add_product_step2')
                elif listing_type == 'bulk':
                    return redirect('vendorapp:bulk_product_upload')
            elif approval and not approval.is_approved:
                if approval.requested_at + timedelta(days=30) > timezone.now():
                    messages.warning(request, f"Please wait 30 days before requesting approval for {brand.brand_name} again.")
                else:
                    return redirect('vendorapp:request_brand_approval', brand_id=brand.id)
            else:
                return redirect('vendorapp:request_brand_approval', brand_id=brand.id)
    
    brand_status = {}
    for brand in brands:
        approval = brand_approvals.filter(brand=brand).first()
        if approval:
            if approval.is_approved:
                status = "Approved"
            elif approval.requested_at + timedelta(days=30) > timezone.now():
                status = "Waiting"
            else:
                status = "Request Again"
        else:
            status = "Not Requested"
        brand_status[str(brand.id)] = status  # Convert brand.id to string

    brand_status_json = json.dumps(brand_status)
    # brand_status_json = json.dumps({"1": "Approved"})

    context = {
        'brands': brands,
        'brand_status': brand_status_json,
    }
    return render(request, 'vendorapp/add_product_step1.html', context)

@login_required
@user_type_required('vendor')
def request_brand_approval(request, brand_id):
    brand = get_object_or_404(Brand, id=brand_id)
    vendor = request.user.vendorprofile

    if request.method == 'POST':
        form = BrandApprovalRequestForm(request.POST)
        if form.is_valid():
            approval_request = form.save(commit=False)
            approval_request.vendor = vendor
            approval_request.brand = brand
            approval_request.save()
            messages.success(request, f"Approval request for {brand.brand_name} has been submitted.")
            return redirect('vendorapp:vendor_dashboard')
    else:
        form = BrandApprovalRequestForm(initial={'brand': brand})

    return render(request, 'vendorapp/request_brand_approval.html', {'form': form, 'brand': brand})

@login_required
@user_type_required('vendor')
def add_product_step2(request):
    selected_brand_id = request.session.get('selected_brand_id')
    brand = get_object_or_404(Brand, id=selected_brand_id)
    vendor = request.user.vendorprofile

    if request.method == 'POST':
        base_watch_form = BaseWatchForm(request.POST, request.FILES)
        details_form = WatchDetailsForm(request.POST)
        materials_form = WatchMaterialsForm(request.POST)

        if all([base_watch_form.is_valid(), details_form.is_valid(), materials_form.is_valid()]):
            base_watch = base_watch_form.save(commit=False)
            base_watch.vendor = vendor
            base_watch.brand = brand

            # Handle primary image upload and color extraction
            if 'primary_image' in request.FILES:
                primary_image = request.FILES['primary_image']
                base_watch.primary_image = primary_image
                
                # Save the image temporarily
                temp_path = default_storage.save('temp_image.jpg', ContentFile(primary_image.read()))
                full_temp_path = os.path.join(settings.MEDIA_ROOT, temp_path)
                
                # Extract dominant color
                dominant_color = get_dominant_color(full_temp_path)
                base_watch.dominant_color = dominant_color
                
                # Remove the temporary file
                os.remove(full_temp_path)

            base_watch.save()
            base_watch_form.save_m2m()

            # Generate image feature
            base_watch.generate_image_feature()

            details = details_form.save(commit=False)
            details.base_watch = base_watch
            details.save()

            materials = materials_form.save(commit=False)
            materials.base_watch = base_watch
            materials.save()

            # Handle multiple image uploads
            images = request.FILES.getlist('images')
            for image in images:
                WatchImage.objects.create(base_watch=base_watch, image=image)

            messages.success(request, f"Product {base_watch.model_name} has been added successfully.")
            return redirect('vendorapp:product_list')
        else:
            # If there are errors, we'll render the form again with error messages
            context = {
                'base_watch_form': base_watch_form,
                'details_form': details_form,
                'materials_form': materials_form,
                'brand': brand,
            }
            return render(request, 'vendorapp/add_product.html', context)
    else:
        base_watch_form = BaseWatchForm()
        details_form = WatchDetailsForm()
        materials_form = WatchMaterialsForm()

    context = {
        'base_watch_form': base_watch_form,
        'details_form': details_form,
        'materials_form': materials_form,
        'brand': brand,
    }
    return render(request, 'vendorapp/add_product.html', context)
    
def get_dominant_color(image_path, k=4):
    img = cv2.imread(image_path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img = img.reshape((img.shape[0] * img.shape[1], 3))
    
    kmeans = KMeans(n_clusters=k)
    kmeans.fit(img)
    
    colors = kmeans.cluster_centers_
    labels = kmeans.labels_
    label_counts = np.bincount(labels)
    
    dominant_color = colors[label_counts.argmax()]
    hex_color = '#{:02x}{:02x}{:02x}'.format(int(dominant_color[0]), int(dominant_color[1]), int(dominant_color[2]))
    
    return hex_color

from django.core.exceptions import ObjectDoesNotExist, ValidationError

@login_required
@user_type_required('vendor')
def bulk_product_upload(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            selected_brand_id = request.session.get('selected_brand_id')
            if not selected_brand_id:
                messages.error(request, "No brand selected. Please select a brand before uploading products.")
                return redirect('vendorapp:add_product_step1')
            
            brand = Brand.objects.get(id=selected_brand_id)
            
            df = pd.read_excel(excel_file, sheet_name='Product Listing')
            
            # Check if required columns are present
            required_columns = ['Model Name', 'MRP', 'Selling Price', 'Available Stock']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"The following required columns are missing: {', '.join(missing_columns)}")
            
            uploaded_products = []
            for index, row in df.iterrows():
                base_watch_data = {
                    'vendor': request.user.vendorprofile,
                    'brand': brand,
                    'is_active': False,
                    'qa_status': False
                }

                # Required fields
                required_fields = [
                    ('model_name', 'Model Name'),
                    ('base_price', 'MRP'),
                    ('selling_price', 'Selling Price'),
                    ('available_stock', 'Available Stock')
                ]
                for field, column in required_fields:
                    if pd.notna(row[column]):
                        base_watch_data[field] = row[column]
                    else:
                        raise ValueError(f"Required field '{column}' is missing for product at row {index + 2}.")

                # Optional fields
                optional_fields = [
                    ('category', 'Category', Category, 'name'),
                    ('collection', 'Collection', Collection, 'name'),
                    ('watch_type', 'Watch Type', WatchType, 'type_name'),
                    ('gender', 'Gender', None, None),
                    ('description', 'Description', None, None),
                    ('color', 'Color', None, None),
                    ('style_code', 'Style Code', None, None),
                    ('net_quantity', 'Net Quantity', None, None),
                    ('function_display', 'Function Display', None, None)
                ]
                for field, column, model, lookup_field in optional_fields:
                    if column in df.columns and pd.notna(row[column]):
                        try:
                            if model:
                                lookup_kwargs = {lookup_field: row[column]}
                                base_watch_data[field] = model.objects.get(**lookup_kwargs)
                            else:
                                base_watch_data[field] = row[column]
                        except (ObjectDoesNotExist, ValidationError):
                            # If the field is invalid or doesn't exist, skip it
                            continue

                base_watch = BaseWatch(**base_watch_data)
                base_watch.total_stock = base_watch.available_stock
                base_watch.save()

                # Create WatchDetails instance
                watch_details_data = {}
                detail_fields = [
                    'case_size', 'water_resistance', 'water_resistance_depth', 'series',
                    'occasion', 'strap_color', 'strap_type', 'dial_color', 'warranty_period'
                ]
                for field in detail_fields:
                    if field in df.columns and pd.notna(row[field]):
                        try:
                            watch_details_data[field] = row[field]
                        except ValidationError:
                            # If the field is invalid, skip it
                            continue

                WatchDetails.objects.create(base_watch=base_watch, **watch_details_data)

                # Create WatchMaterials instance
                watch_materials_data = {}
                material_fields = ['strap_material', 'glass_material', 'case_material']
                for field in material_fields:
                    if field in df.columns and pd.notna(row[field]):
                        try:
                            watch_materials_data[field] = Material.objects.get(name=row[field])
                        except ObjectDoesNotExist:
                            # If the material doesn't exist, skip it
                            continue

                WatchMaterials.objects.create(base_watch=base_watch, **watch_materials_data)

                # Add features
                if 'Features' in df.columns and pd.notna(row['Features']):
                    features = row['Features'].split(',')
                    for feature_name in features:
                        feature_name = feature_name.strip()
                        if feature_name:
                            try:
                                feature = Feature.objects.get(name=feature_name)
                                base_watch.features.add(feature)
                            except ObjectDoesNotExist:
                                # If the feature doesn't exist, skip it
                                continue

                uploaded_products.append(base_watch.id)

            request.session['uploaded_products'] = uploaded_products
            messages.success(request, f'{len(uploaded_products)} products uploaded successfully!')
            return redirect('vendorapp:upload_product_images')
        except Exception as e:
            messages.error(request, f'Error uploading products: {str(e)}')
    
    return render(request, 'vendorapp/bulk_product_upload.html')

@login_required
@user_type_required('vendor')
def upload_product_images(request):
    uploaded_products = request.session.get('uploaded_products', [])
    products = BaseWatch.objects.filter(id__in=uploaded_products)
    
    if request.method == 'POST':
        for product in products:
            primary_image = request.FILES.get(f'primary_image_{product.id}')
            if primary_image:
                product.primary_image = primary_image
                product.save()
            
            additional_images = request.FILES.getlist(f'additional_images_{product.id}')
            for image in additional_images:
                WatchImage.objects.create(base_watch=product, image=image)
        
        messages.success(request, 'Images uploaded successfully!')
        return redirect('vendorapp:product_list')
    
    context = {
        'products': products
    }
    return render(request, 'vendorapp/upload_product_images.html', context)



@login_required
@user_type_required('vendor')
def download_template(request):
    wb = Workbook()
    
    # Guidelines sheet
    guidelines = wb.active
    guidelines.title = "Guidelines"
    guidelines.column_dimensions['A'].width = 100
    
    guidelines.merge_cells('A1:C1')
    cell = guidelines['A1']
    cell.value = "Guidelines for Bulk Product Upload"
    cell.font = Font(size=16, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    guidelines.append([])
    guidelines.append(["1. Fill out the 'Product Listing' sheet with your product information."])
    guidelines.append(["2. Ensure all required fields (marked in RED) are filled."])
    guidelines.append(["3. Optional fields (marked in GREEN) can be left blank or filled as needed."])
    guidelines.append(["4. Use dropdown menus where available to ensure accurate data entry."])
    guidelines.append(["5. For multiple features, separate them with commas."])
    guidelines.append(["6. Dates should be in YYYY-MM-DD format."])
    guidelines.append(["7. Ensure numeric fields contain only numbers."])
    guidelines.append(["8. Case Size should be in the format: 23 X 34 X 45 (length X width X height in mm)"])
    guidelines.append(["9. MRP (Market Retail Price) should be greater than or equal to the Selling Price."])
    
    for row in range(2, 11):
        guidelines[f'A{row}'].font = Font(size=12)
    
    guidelines.append([])
    guidelines.merge_cells('A12:C12')
    cell = guidelines['A12']
    cell.value = "Color Code:"
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    guidelines.append(["RED - Required fields"])
    guidelines['A13'].font = Font(size=12, color="FF0000")
    guidelines['A13'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    guidelines.append(["GREEN - Optional fields"])
    guidelines['A14'].font = Font(size=12, color="008000")
    guidelines['A14'].fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    
    # Product Listing sheet
    product_listing = wb.create_sheet("Product Listing")
    headers = [
        ('Model Name', True), ('Category', True), ('Collection', True), ('Watch Type', False),
        ('Gender', True), ('MRP', True), ('Selling Price', True), ('Description', True),
        ('Available Stock', True), ('Color', True), ('Style Code', True),
        ('Net Quantity', True), ('Function Display', False), ('Case Size', True),
        ('Water Resistance', False), ('Water Resistance Depth', False), ('Series', False),
        ('Occasion', False), ('Strap Color', True), ('Strap Type', True), ('Dial Color', True),
        ('Warranty Period', True), ('Strap Material', True), ('Glass Material', True),
        ('Case Material', True), ('Features', False)
    ]
    
    for col, (header, required) in enumerate(headers, start=1):
        cell = product_listing.cell(row=1, column=col, value=header)
        cell.font = Font(size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF0000" if required else "008000", end_color="FF0000" if required else "008000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        product_listing.column_dimensions[cell.column_letter].width = 20
    
    # Add dropdowns
    collections = [col.name for col in Collection.objects.all()]
    categories = [cat.name for cat in Category.objects.all()]
    watch_types = [wt.type_name for wt in WatchType.objects.all()]
    genders = ['Men', 'Women', 'Unisex']
    colors = ['Black', 'White', 'Silver', 'Gold', 'Blue', 'Red', 'Green', 'Yellow', 'Brown', 'Other']
    function_displays = ['Analog', 'Digital', 'Analog-Digital']
    strap_types = ['Bracelet', 'Leather', 'Rubber', 'Fabric', 'Other']
    materials = [mat.name for mat in Material.objects.all()]
    
    dv_collection = DataValidation(type="list", formula1=f'"{",".join(collections)}"', allow_blank=True)
    dv_category = DataValidation(type="list", formula1=f'"{",".join(categories)}"', allow_blank=True)
    dv_watch_type = DataValidation(type="list", formula1=f'"{",".join(watch_types)}"', allow_blank=True)
    dv_gender = DataValidation(type="list", formula1=f'"{",".join(genders)}"', allow_blank=True)
    dv_color = DataValidation(type="list", formula1=f'"{",".join(colors)}"', allow_blank=True)
    dv_function_display = DataValidation(type="list", formula1=f'"{",".join(function_displays)}"', allow_blank=True)
    dv_strap_type = DataValidation(type="list", formula1=f'"{",".join(strap_types)}"', allow_blank=True)
    dv_material = DataValidation(type="list", formula1=f'"{",".join(materials)}"', allow_blank=True)
    
    product_listing.add_data_validation(dv_category)
    product_listing.add_data_validation(dv_watch_type)
    product_listing.add_data_validation(dv_gender)
    product_listing.add_data_validation(dv_color)
    product_listing.add_data_validation(dv_function_display)
    product_listing.add_data_validation(dv_strap_type)
    product_listing.add_data_validation(dv_material)
    
    # Add data validations to the Product Listing sheet
    dv_category.add(f'B2:B1048576')
    dv_collection.add(f'C2:C1048576')
    dv_watch_type.add(f'D2:D1048576')
    dv_gender.add(f'E2:E1048576')
    dv_color.add(f'J2:J1048576')
    dv_function_display.add(f'M2:M1048576')
    dv_strap_type.add(f'T2:T1048576')
    dv_material.add(f'W2:Y1048576')  # For Strap, Glass, and Case materials
    
    # Save the workbook
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=product_upload_template.xlsx'
    return response



@login_required
@user_type_required('vendor')
def analysis_view(request):
    # Get the current vendor
    vendor = request.user.vendorprofile

    # Get the selected time range from the request, default to 28 days
    time_range = request.GET.get('time_range', '28')
    
    # Define time ranges
    time_ranges = {
        '7': timedelta(days=7),
        '28': timedelta(days=28),
        '90': timedelta(days=90),
        '365': timedelta(days=365),
    }
    
    # Handle custom date range
    if time_range == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and end_date:
            start_date = make_aware(timezone.datetime.strptime(start_date, '%Y-%m-%d'))
            end_date = make_aware(timezone.datetime.strptime(end_date, '%Y-%m-%d')).replace(hour=23, minute=59, second=59)
        else:
            # Default to last 28 days if custom dates are not provided
            end_date = timezone.now()
            start_date = end_date - timedelta(days=27)
    elif time_range in time_ranges:
        end_date = timezone.now()
        start_date = end_date - time_ranges[time_range]
    else:
        # Handle year and month selections
        year = int(time_range[:4])
        if len(time_range) == 4:  # Year selection
            start_date = make_aware(timezone.datetime(year, 1, 1))
            end_date = make_aware(timezone.datetime(year, 12, 31, 23, 59, 59))
        else:  # Month selection
            month = int(time_range[5:])
            start_date = make_aware(timezone.datetime(year, month, 1))
            _, last_day = calendar.monthrange(year, month)
            end_date = make_aware(timezone.datetime(year, month, last_day, 23, 59, 59))

    # Total watches sold (only for this vendor's products)
    total_watches = OrderItem.objects.filter(
        order__created_at__gte=start_date,
        order__created_at__lte=end_date,
        watch__vendor=vendor
    ).aggregate(Sum('quantity'))['quantity__sum'] or 0

    # Calculate revenue (only for this vendor's products)
    total_revenue = OrderItem.objects.filter(
        order__created_at__gte=start_date,
        order__created_at__lte=end_date,
        watch__vendor=vendor
    ).aggregate(total=Sum(F('quantity') * F('price')))['total'] or 0

    # Compare with previous period (only for this vendor's products)
    previous_start = start_date - (end_date - start_date)
    previous_revenue = OrderItem.objects.filter(
        order__created_at__gte=previous_start,
        order__created_at__lt=start_date,
        watch__vendor=vendor
    ).aggregate(total=Sum(F('quantity') * F('price')))['total'] or 0

    if previous_revenue > 0:
        revenue_comparison = round(((total_revenue - previous_revenue) / previous_revenue) * 100, 2)
    elif total_revenue > 0:
        revenue_comparison = 100  # 100% increase if there was no previous revenue but there is current revenue
    else:
        revenue_comparison = 0  # No change if both periods have zero revenue

    # New customers (only for this vendor)
    new_customers = Order.objects.filter(
        created_at__gte=start_date,
        created_at__lte=end_date,
        items__watch__vendor=vendor
    ).values('user').distinct().count()

    previous_new_customers = Order.objects.filter(
        created_at__gte=previous_start,
        created_at__lt=start_date,
        items__watch__vendor=vendor
    ).values('user').distinct().count()

    customer_comparison = round(((new_customers - previous_new_customers) / previous_new_customers) * 100 if previous_new_customers else 0, 2)
    
    # Create a list of all dates in the range
    all_dates = [start_date.date() + timedelta(days=x) for x in range((end_date.date() - start_date.date()).days + 1)]

    # Get the sales data (only for this vendor's products)
    sales_data = OrderItem.objects.filter(
        order__created_at__date__gte=start_date.date(),
        order__created_at__date__lte=end_date.date(),
        watch__vendor=vendor
    ).annotate(
        date=TruncDate('order__created_at')
    ).values('date').annotate(
        sales=Sum('quantity')
    ).order_by('date')

    # Convert to a dictionary for easy lookup
    sales_dict = {item['date']: item['sales'] for item in sales_data}

    # Create the final data, including zero sales days
    daily_sales = [{'date': date, 'sales': sales_dict.get(date, 0)} for date in all_dates]

    date_labels = [item['date'].strftime('%d %b') for item in daily_sales]
    daily_sales_data = [item['sales'] for item in daily_sales]

    # Realtime data (last 48 hours, only for this vendor's products)
    last_48_hours = timezone.now() - timedelta(hours=48)
    realtime_sales = OrderItem.objects.filter(
        order__created_at__gte=last_48_hours,
        watch__vendor=vendor
    ).annotate(hour=TruncHour('order__created_at'))\
        .values('hour')\
        .annotate(sales=Sum('quantity'))\
        .order_by('hour')

    realtime_labels = [(timezone.now() - timedelta(hours=i)).strftime('%H:%M') for i in range(48, 0, -1)]
    realtime_data = [0] * 48

    for sale in realtime_sales:
        hours_ago = int((timezone.now() - sale['hour']).total_seconds() / 3600)
        if hours_ago < 48:
            realtime_data[47 - hours_ago] = sale['sales']

    recent_sales = sum(realtime_data)

    # Top selling watches (only for this vendor)
    top_watches = BaseWatch.objects.filter(vendor=vendor).annotate(
        sales_count=Sum('orderitem__quantity')
    ).filter(
        orderitem__order__created_at__gte=start_date,
        orderitem__order__created_at__lte=end_date,
        primary_image__isnull=False
    ).order_by('-sales_count')[:5]

    # Total customers (only for this vendor)
    total_customers = Order.objects.filter(
        items__watch__vendor=vendor
    ).values('user').distinct().count()

    # Prepare date range options for the template
    current_year = timezone.now().year
    date_range_options = [
        {'value': '7', 'label': 'Last 7 days'},
        {'value': '28', 'label': 'Last 28 days'},
        {'value': '90', 'label': 'Last 90 days'},
        {'value': '365', 'label': 'Last 365 days'},
        {'value': str(current_year), 'label': str(current_year)},
        {'value': str(current_year - 1), 'label': str(current_year - 1)},
    ]

    # Add current year's months
    current_date = timezone.now()
    current_year = current_date.year
    current_month = current_date.month

    for month_offset in range(2, -1, -1):
        target_date = current_date - relativedelta(months=month_offset)
        date_range_options.append({
            'value': f"{target_date.year}-{target_date.month:02d}",
            'label': f"{calendar.month_name[target_date.month]} {target_date.year}"
        })

    context = {
        'total_watches': total_watches,
        'total_revenue': total_revenue,
        'revenue_comparison': revenue_comparison,
        'new_customers': new_customers,
        'customer_comparison': customer_comparison,
        'start_date': start_date,
        'end_date': end_date,
        'date_labels': json.dumps(date_labels),
        'daily_sales': json.dumps(daily_sales_data),
        'recent_sales': recent_sales,
        'realtime_labels': json.dumps(realtime_labels),
        'realtime_data': json.dumps(realtime_data),
        'top_watches': top_watches,
        'total_customers': total_customers,
        'date_range_options': date_range_options,
        'selected_range': time_range,
    }

    return render(request, 'vendorapp/analysis.html', context)